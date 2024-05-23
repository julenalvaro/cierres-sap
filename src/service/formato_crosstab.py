# PATH: src/service/formato_crosstab.py

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def cargar_umb(bom, modelo):
    """ Devuelve un diccionario de las cantidades UMB por componente para el modelo dado. """
    bom_filtrado = bom[bom['Modelo'] == modelo]
    return bom_filtrado.set_index('Nº componentes')['Ctd.componente (UMB)'].to_dict()

def agregar_cantidad_bom_header(ws, bom, modelo):
    """ Ajusta los encabezados de material y añade una fila para Cantidad_BOM. """
    umb_dict = cargar_umb(bom, modelo)
    ws.cell(row=1, column=1, value="Material").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ws.cell(row=2, column=1, value="Cantidad_BOM").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=2, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col in range(2, ws.max_column + 1):
        material = ws.cell(row=1, column=col).value.split(' (')[0]  # Eliminar cantidades anteriores si están presentes
        cantidad_bom = umb_dict.get(material, 'N/A')
        ws.cell(row=1, column=col, value=material)
        ws.cell(row=2, column=col, value=cantidad_bom)
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

def format_crosstabs(ws, bom, modelo):
    umb_dict = cargar_umb(bom, modelo)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.value = 0 if cell.value is None else cell.value
            material = ws.cell(row=1, column=cell.column).value
            umb = umb_dict.get(material, 0)
            if cell.value < umb:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell.value > umb:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif cell.value == umb:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Colorear la columna 1 de la fila 3 hasta max_row con gris clarito
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    # Congelar las dos primeras filas
    ws.freeze_panes = ws['B3']

    # Ajustar el ancho de las columnas al contenido
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
