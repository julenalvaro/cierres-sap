# PATH: src/service/formato_arbol_correcciones.py

from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

def agregar_enlace_arbol(ws, modelos):
    # Encontrar la columna que contiene el encabezado "Modelo"
    modelo_col = None
    for cell in ws[1]:
        if cell.value == "Modelo":
            modelo_col = cell.column_letter
            break

    if modelo_col is None:
        raise ValueError("No se encontró una columna con el encabezado 'Modelo'")

    # Agregar enlaces a las celdas en la columna "Modelo"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[ws[modelo_col + '1'].col_idx - 1]  # Obtener la celda en la columna "Modelo" de la fila actual
        if cell.value in modelos:
            modelo = cell.value
            cell.hyperlink = Hyperlink(ref=f'{modelo_col}{cell.row}', target=f'#\'{modelo}\'!A1', tooltip=f'Ir a {modelo}')
            cell.font = Font(color='0000FF', underline='single')

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def format_arbol_correcciones(arbol_ws):
    # Filtrar las filas que tienen valores en blanco o cadena vacía en "margen_ordenes"
    col_margen_ordenes = None
    for cell in arbol_ws[1]:  # Buscar columna "margen_ordenes" en la primera fila
        if cell.value == "margen_ordenes":
            col_margen_ordenes = cell.column_letter
            break
    
    if col_margen_ordenes:
        rows_to_delete = []
        for row in arbol_ws.iter_rows(min_row=2):  # Omite el encabezado
            cell_value = row[arbol_ws[col_margen_ordenes + '1'].column - 1].value
            if cell_value is None or cell_value == "":
                rows_to_delete.append(row[0].row)
        
        for row in reversed(rows_to_delete):  # Elimina filas de abajo hacia arriba
            arbol_ws.delete_rows(row)
    
    # Formatear como tabla
    tab = Table(displayName="ArbolCorrecciones", ref=arbol_ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    arbol_ws.add_table(tab)

    # Ajustar el ancho de las columnas al contenido, excluyendo el encabezado
    columns_to_exclude = ["pos_estructura", "Nivel explosión", "margen_ordenes", "stocks"]
    for col in arbol_ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtén la letra de la columna
        column_header = col[0].value
        if column_header not in columns_to_exclude:
            for cell in col[1:]:  # Omite el encabezado
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
        else:
            adjusted_width = len(column_header) + 2
        arbol_ws.column_dimensions[column].width = adjusted_width

    # Añadir comentarios a los encabezados
    for cell in arbol_ws[1]:  # Primera fila (encabezados)
        cell.comment = Comment(text=cell.value, author="AutoFormat")

    # Centrar el contenido de todas las celdas, excepto pos_estructura y Nivel explosión
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    for row in arbol_ws.iter_rows():
        for cell in row:
            if cell.row == 1:  # Saltar encabezados
                continue
            if cell.column_letter in [col[0].column_letter for col in arbol_ws.columns if col[0].value in ["pos_estructura", "Nivel explosión"]]:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

    # Formato condicional para la columna "margen_ordenes"
    if col_margen_ordenes:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        if arbol_ws.max_row > 1:  # Asegurarse de que hay filas para aplicar el formato condicional
            arbol_ws.conditional_formatting.add(f"{col_margen_ordenes}2:{col_margen_ordenes}{arbol_ws.max_row}",
                                                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))
            arbol_ws.conditional_formatting.add(f"{col_margen_ordenes}2:{col_margen_ordenes}{arbol_ws.max_row}",
                                                CellIsRule(operator="greaterThan", formula=["0"], fill=yellow_fill))
            arbol_ws.conditional_formatting.add(f"{col_margen_ordenes}2:{col_margen_ordenes}{arbol_ws.max_row}",
                                                CellIsRule(operator="equal", formula=["0"], fill=green_fill))

    # Congelar paneles para que los encabezados siempre sean visibles
    arbol_ws.freeze_panes = arbol_ws['A2']