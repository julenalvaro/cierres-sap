from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

from src.service.generar_crosstab_modelo_materiales import cargar_datos, preparar_datos_coois, generar_crosstab_modelo_materiales
from src.service.formato_crosstab import format_crosstabs, agregar_cantidad_bom_header, agregar_enlace_indice, formato_indice

def agregar_enlace_indice_hoja(ws):
    """Agrega un enlace a la hoja de índice en la primera celda de la hoja."""
    cell = ws.cell(row=1, column=1, value="Volver al índice")
    link = f"#'Índice'!A1"
    cell.hyperlink = Hyperlink(ref="", location=link, display="Volver al índice")
    cell.font = Font(color="0000FF", underline="single")
    cell.alignment = Alignment(horizontal="center")

def generar_excel_crosstabs_acabados(archivo, sheet_bom, sheet_coois):
    bom, coois = cargar_datos(archivo, sheet_bom, sheet_coois)
    coois = preparar_datos_coois(coois)

    wb = Workbook()
    wb.remove(wb.active)  # Elimina la pestaña predeterminada
    
    # Crear pestaña de índice
    index_sheet = wb.create_sheet(title="Índice", index=0)
    index_sheet.append(["Modelo"])

    unique_modelos = sorted(bom['Modelo'].unique())

    for i, modelo in enumerate(unique_modelos, start=2):
        ws = wb.create_sheet(title=modelo)
        crosstab = generar_crosstab_modelo_materiales(bom, coois, modelo)
        for r in dataframe_to_rows(crosstab, index=True, header=True):
            ws.append(r)
        agregar_cantidad_bom_header(ws, bom, modelo)
        format_crosstabs(ws, bom, modelo)

        # Agregar enlace al índice y darle formato
        agregar_enlace_indice(index_sheet, modelo, i)
        agregar_enlace_indice_hoja(ws)

    # Aplicar formato al índice
    formato_indice(index_sheet)

    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f'./results/crosstabs_acabados/crosstabs_materiales_acabados_{timestamp}.xlsx'
    wb.save(filename)
    return filename
