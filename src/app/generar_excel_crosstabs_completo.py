# PATH: src/app/generar_excel_crosstabs_completo.py

import os
import traceback
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config.config import obtener_configuracion
from src.service.generar_crosstab_modelo_materiales import cargar_datos_maestros, cargar_coois, cargar_stocks, transformar_coois, transformar_stocks, transformar_fabricacion_real, generar_crosstab_modelo_materiales
from src.service.formato_crosstab import format_crosstabs, agregar_cantidad_bom_header, formato_indice, agregar_enlace_indice, agregar_enlace_indice_hoja, guardar_excel
from src.service.transformar_bom_a_arbol_correcciones import transformar_bom_a_arbol_correcciones_EA, transformar_bom_a_arbol_correcciones_EB
from src.service.formato_arbol_correcciones import agregar_enlace_arbol, format_arbol_correcciones

# #################debug
# from datetime import datetime
# #################debug

def generar_excel_crosstabs_completo(archivo, sheet_bom_ea, sheet_bom_eb, sheet_coois, sheet_stocks, sheet_fabricacion_real_ea, sheet_fabricacion_real_eb):
    config = obtener_configuracion()
    dir_crosstabs = os.path.join(config.DIR_BASE, "informes_crosstab")

    try:
        print('Cargando datos...')
        
        bom_ea, bom_eb, fabricacion_real_ea, fabricacion_real_eb = cargar_datos_maestros(
            archivo, 
            sheet_bom_ea, 
            sheet_bom_eb, 
            sheet_coois, 
            sheet_stocks,
            sheet_fabricacion_real_ea,
            sheet_fabricacion_real_eb
        )

        print('Preparando datos...')

        # Transformaciones primarias tablas, traducidas del código M 
        
        coois_ea, coois_eb = transformar_coois(coois)
        stocks = transformar_stocks(stocks)
        fabricacion_real_eb = transformar_fabricacion_real(fabricacion_real_eb)
        fabricacion_real_ea = transformar_fabricacion_real(fabricacion_real_ea)

        # #################debug
        # # Generar timestamp para los archivos
        # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # # Guardar los DataFrames transformados para debugging
        # debug_dir = './archivos/debug'
        # os.makedirs(debug_dir, exist_ok=True)
        
        # coois_ea.to_excel(os.path.join(debug_dir, f'coois_ea_{timestamp}.xlsx'))
        # coois_eb.to_excel(os.path.join(debug_dir, f'coois_eb_{timestamp}.xlsx'))
        # stocks.to_excel(os.path.join(debug_dir, f'stocks_{timestamp}.xlsx'))
        # fabricacion_real_ea.to_excel(os.path.join(debug_dir, f'fabricacion_real_ea_{timestamp}.xlsx'))
        # fabricacion_real_eb.to_excel(os.path.join(debug_dir, f'fabricacion_real_eb_{timestamp}.xlsx'))
        # #################debug

        results = []

        for bom, coois, fabricacion_real, subset_name in [(bom_ea, coois_ea, fabricacion_real_ea, 'EA'), (bom_eb, coois_eb, fabricacion_real_eb, 'EB')]:
            try:
                print(f'Generando archivo Excel para {subset_name}...')
                wb = Workbook()
                wb.remove(wb.active)  # Elimina la pestaña predeterminada

                arbol_ws = wb.create_sheet(title=f'arbol_correcciones_{subset_name}', index=0)
                index_sheet = wb.create_sheet(title="Indice", index=1)
                index_sheet.append(["Modelo"])
                unique_modelos = sorted(bom['Modelo'].unique())
                for modelo in unique_modelos:
                    index_sheet.append([modelo]) 

                # Transformar y agregar el árbol de correcciones primero para vincular crosstabs
                print('Transformando árbol...')

                if subset_name == 'EA':
                    arbol_correcciones = transformar_bom_a_arbol_correcciones_EA(bom, coois, fabricacion_real, stocks)
                elif subset_name == 'EB':
                    arbol_correcciones = transformar_bom_a_arbol_correcciones_EB(bom, coois, fabricacion_real, stocks)
                arbol_correcciones = arbol_correcciones.astype('object')
                arbol_correcciones.fillna('', inplace=True)
                
                # ######debug
                # arbol_correcciones.to_excel(os.path.join(debug_dir, f'arbol_correcciones_{subset_name}_{timestamp}.xlsx'))
                # ######debug

                for row in dataframe_to_rows(arbol_correcciones, index=False, header=True):
                    arbol_ws.append(row)
                
                # Generar crosstabs modelos
                for i, modelo in enumerate(unique_modelos, start=2):
                    ws = wb.create_sheet(title=modelo, index=i + 2)
                    crosstab = generar_crosstab_modelo_materiales(bom, coois, modelo)
                    for row in dataframe_to_rows(crosstab, index=True, header=True):
                        ws.append(row)
                    agregar_cantidad_bom_header(ws, bom, modelo)
                    format_crosstabs(ws, bom, modelo)

                    # Agregar enlace desde crosstabs al árbol de correcciones y no al índice
                    agregar_enlace_indice_hoja(ws, 'arbol_correcciones_' + subset_name)
                    agregar_enlace_indice(index_sheet, modelo, i + 1)

                formato_indice(index_sheet)
                format_arbol_correcciones(arbol_ws)
                agregar_enlace_arbol(arbol_ws, unique_modelos)

                # Reordenar hojas: árbol de correcciones primero, luego índice, luego crosstabs
                wb._sheets.sort(key=lambda sheet: sheet.title not in [f'arbol_correcciones_{subset_name}', "Índice"])

                filename = guardar_excel(wb, dir_crosstabs, subset_name)
                results.append(filename)
                print(f'Archivo generado para {subset_name}: {filename}')
            except Exception as e:
                print(f'Ocurrió un error generando el archivo para {subset_name}: {e}')
                traceback.print_exc()  # Imprimir el rastreo completo del error
                results.append(None)

        return results
    except Exception as e:
        print(f'Ocurrió un error en la función principal: {e}')
        traceback.print_exc()  # Imprimir el rastreo completo del error
        return None, None

